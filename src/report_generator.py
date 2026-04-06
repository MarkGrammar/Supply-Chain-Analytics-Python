"""
Excel report generator using openpyxl.

Produces a professional 5-sheet workbook:
  1. Executive Summary  — portfolio KPI scorecard + charts
  2. Supplier Performance — per-supplier KPI table + chart
  3. Monthly Trends     — time-series table + combo chart
  4. At-Risk Open Orders — risk-flagged orders with color coding
  5. Raw Data           — full enriched dataset with auto-filter
"""

import os
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.layout import Layout

import config

# ── Colour palette ────────────────────────────────────────────────────────────
C_DARK_BLUE  = "1F3864"
C_MID_BLUE   = "2E75B6"
C_LIGHT_BLUE = "BDD7EE"
C_WHITE      = "FFFFFF"
C_LIGHT_GREY = "F2F2F2"
C_GREEN      = "70AD47"
C_AMBER      = "FFC000"
C_RED        = "FF0000"
C_DARK_RED   = "C00000"
C_LIGHT_RED  = "FFCCCC"
C_LIGHT_AMB  = "FFE699"
C_LIGHT_GRN  = "E2EFDA"

# Risk level colours
RISK_FILLS = {
    "CRITICAL": PatternFill("solid", fgColor="C00000"),
    "HIGH":     PatternFill("solid", fgColor="FF4500"),
    "MEDIUM":   PatternFill("solid", fgColor="FFC000"),
    "LOW":      PatternFill("solid", fgColor="70AD47"),
}
RISK_FONTS = {
    "CRITICAL": Font(color=C_WHITE, bold=True),
    "HIGH":     Font(color=C_WHITE, bold=True),
    "MEDIUM":   Font(color="000000", bold=True),
    "LOW":      Font(color=C_WHITE),
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def _header_fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)

def _bold_font(color: str = "000000", size: int = 11) -> Font:
    return Font(bold=True, color=color, size=size)

def _thin_border() -> Border:
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _write_df_to_sheet(ws, df: pd.DataFrame, start_row: int = 1, start_col: int = 1,
                        header_color: str = C_DARK_BLUE):
    """Write a DataFrame to a worksheet with styled headers."""
    for col_idx, col_name in enumerate(df.columns, start=start_col):
        cell = ws.cell(row=start_row, column=col_idx, value=str(col_name).replace("_", " ").title())
        cell.font = Font(bold=True, color=C_WHITE)
        cell.fill = _header_fill(header_color)
        cell.alignment = _center()
        cell.border = _thin_border()

    for row_idx, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for col_idx, value in enumerate(row, start=start_col):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = _thin_border()
            cell.alignment = Alignment(vertical="center")
            # Alternate row shading
            if (row_idx - start_row) % 2 == 0:
                cell.fill = _header_fill(C_LIGHT_GREY)

def _auto_col_width(ws, min_width: int = 10, max_width: int = 40):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


# ── Sheet 1: Executive Summary ────────────────────────────────────────────────

def _sheet_executive_summary(wb: Workbook, pipeline_results: dict, at_risk: pd.DataFrame):
    ws = wb.create_sheet("Executive Summary")
    ws.sheet_view.showGridLines = False

    p = pipeline_results["portfolio_kpis"].iloc[0]

    # Title banner
    ws.merge_cells("A1:H2")
    title_cell = ws["A1"]
    title_cell.value = "Supply Chain Purchase Order Analytics — Executive Summary"
    title_cell.font = Font(bold=True, color=C_WHITE, size=16)
    title_cell.fill = _header_fill(C_DARK_BLUE)
    title_cell.alignment = _center()

    ws.merge_cells("A3:H3")
    ts_cell = ws["A3"]
    ts_cell.value = f"Report generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  Data period: {config.SIMULATION_START_DATE} – {config.SIMULATION_END_DATE}"
    ts_cell.font = Font(italic=True, color=C_DARK_BLUE, size=10)
    ts_cell.alignment = Alignment(horizontal="center")
    ts_cell.fill = _header_fill(C_LIGHT_BLUE)

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 18

    # KPI scorecard grid (2 rows × 4 cols)
    kpis = [
        ("Total Orders",           f"{p['total_orders']:,}",          None),
        ("On-Time Delivery Rate",  f"{p['otd_rate']:.1%}",            "otd"),
        ("Avg Lead Time",          f"{p['avg_lead_time_days']:.1f} days", None),
        ("Open Orders",            f"{p['open_orders']:,}",            None),
        ("Total Spend",            f"${p['total_spend']:,.0f}",        None),
        ("Total Delay Cost",       f"${p['total_delay_cost']:,.0f}",   "delay"),
        ("Delay Cost % of Spend",  f"{p['delay_cost_pct_spend']:.2%}", "delay"),
        ("At-Risk Orders",         f"{len(at_risk)}",                  "risk"),
    ]

    row_offset = 5
    for i, (label, value, flag) in enumerate(kpis):
        col = (i % 4) * 2 + 1
        row = row_offset + (i // 4) * 3

        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        lbl = ws.cell(row=row, column=col, value=label)
        lbl.font = _bold_font(C_WHITE)
        lbl.fill = _header_fill(C_MID_BLUE)
        lbl.alignment = _center()
        lbl.border = _thin_border()

        ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+1)
        val = ws.cell(row=row+1, column=col, value=value)
        val.font = Font(bold=True, size=14)
        val.alignment = _center()
        val.border = _thin_border()

        if flag == "otd":
            otd = p["otd_rate"]
            color = C_LIGHT_GRN if otd >= 0.90 else (C_LIGHT_AMB if otd >= 0.75 else C_LIGHT_RED)
            val.fill = _header_fill(color)
        elif flag == "delay":
            pct = p["delay_cost_pct_spend"]
            color = C_LIGHT_GRN if pct < 0.01 else (C_LIGHT_AMB if pct < 0.03 else C_LIGHT_RED)
            val.fill = _header_fill(color)
        elif flag == "risk":
            risk_n = len(at_risk)
            color = C_LIGHT_GRN if risk_n == 0 else (C_LIGHT_AMB if risk_n < 10 else C_LIGHT_RED)
            val.fill = _header_fill(color)
        else:
            val.fill = _header_fill(C_LIGHT_GREY)

        ws.row_dimensions[row].height = 20
        ws.row_dimensions[row+1].height = 28

    # Chart 1: Top 10 suppliers by spend
    sup = pipeline_results["supplier_kpis"].nlargest(10, "total_spend")[["supplier_name", "total_spend"]].copy()
    chart_start_row = row_offset + 7
    _write_df_to_sheet(ws, sup, start_row=chart_start_row, start_col=1, header_color=C_MID_BLUE)

    bar = BarChart()
    bar.type = "bar"
    bar.title = "Top 10 Suppliers by Total Spend"
    bar.y_axis.title = "Spend (USD)"
    bar.x_axis.title = "Supplier"
    bar.style = 10
    bar.width = 18
    bar.height = 12

    data_ref = Reference(ws, min_col=2, min_row=chart_start_row,
                          max_row=chart_start_row + len(sup))
    cats_ref = Reference(ws, min_col=1, min_row=chart_start_row + 1,
                          max_row=chart_start_row + len(sup))
    bar.add_data(data_ref, titles_from_data=True)
    bar.set_categories(cats_ref)
    ws.add_chart(bar, f"E{chart_start_row}")

    # Chart 2: Spend by category (pie)
    cat_spend = pipeline_results["orders"].groupby("product_category")["total_po_value"].sum().reset_index()
    cat_start_row = chart_start_row + len(sup) + 3
    _write_df_to_sheet(ws, cat_spend, start_row=cat_start_row, start_col=1, header_color=C_MID_BLUE)

    pie = PieChart()
    pie.title = "Total Spend by Category"
    pie.style = 10
    pie.width = 14
    pie.height = 10
    pie_data = Reference(ws, min_col=2, min_row=cat_start_row,
                          max_row=cat_start_row + len(cat_spend))
    pie_cats = Reference(ws, min_col=1, min_row=cat_start_row + 1,
                          max_row=cat_start_row + len(cat_spend))
    pie.add_data(pie_data, titles_from_data=True)
    pie.set_categories(pie_cats)
    ws.add_chart(pie, f"E{cat_start_row}")

    for col_idx in range(1, 9):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16


# ── Sheet 2: Supplier Performance ─────────────────────────────────────────────

def _sheet_supplier_performance(wb: Workbook, supplier_kpis: pd.DataFrame):
    ws = wb.create_sheet("Supplier Performance")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:K1")
    ws["A1"].value = "Supplier Performance Dashboard"
    ws["A1"].font = Font(bold=True, color=C_WHITE, size=14)
    ws["A1"].fill = _header_fill(C_DARK_BLUE)
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 28

    display_cols = [
        "supplier_name", "tier", "order_count", "total_spend",
        "otd_rate", "avg_lead_time", "std_lead_time",
        "total_delay_cost", "delay_cost_pct_spend",
        "reliability_score", "flagged_unreliable",
    ]
    sup_display = supplier_kpis[display_cols].copy()
    sup_display = sup_display.sort_values("reliability_score", ascending=False)

    _write_df_to_sheet(ws, sup_display, start_row=3, header_color=C_MID_BLUE)

    # Conditional formatting: reliability_score column (col index 10 → col J)
    rel_col = display_cols.index("reliability_score") + 1
    rel_col_letter = get_column_letter(rel_col)
    data_end = 3 + len(sup_display)

    ws.conditional_formatting.add(
        f"{rel_col_letter}4:{rel_col_letter}{data_end}",
        ColorScaleRule(
            start_type="num", start_value=0.0, start_color="FF0000",
            mid_type="num",   mid_value=0.75,  mid_color="FFC000",
            end_type="num",   end_value=1.0,   end_color="70AD47",
        )
    )

    # Highlight unreliable supplier rows
    flag_col = display_cols.index("flagged_unreliable") + 1
    flag_col_letter = get_column_letter(flag_col)
    for row_idx in range(4, data_end + 1):
        cell = ws.cell(row=row_idx, column=flag_col)
        if cell.value is True:
            for c in range(1, len(display_cols) + 1):
                ws.cell(row=row_idx, column=c).fill = _header_fill("FFCCCC")

    # Chart: reliability scores ranked
    chart_data = sup_display[["supplier_name", "reliability_score"]].reset_index(drop=True)
    chart_row = data_end + 3
    _write_df_to_sheet(ws, chart_data, start_row=chart_row, start_col=1, header_color=C_MID_BLUE)

    bar = BarChart()
    bar.type = "bar"
    bar.title = "Supplier Reliability Scores"
    bar.y_axis.title = "Reliability Score"
    bar.style = 10
    bar.width = 22
    bar.height = 14
    data_ref = Reference(ws, min_col=2, min_row=chart_row,
                          max_row=chart_row + len(chart_data))
    cats_ref = Reference(ws, min_col=1, min_row=chart_row + 1,
                          max_row=chart_row + len(chart_data))
    bar.add_data(data_ref, titles_from_data=True)
    bar.set_categories(cats_ref)
    ws.add_chart(bar, f"D{chart_row}")

    _auto_col_width(ws)
    ws.freeze_panes = "A4"


# ── Sheet 3: Monthly Trends ────────────────────────────────────────────────────

def _sheet_monthly_trends(wb: Workbook, monthly_kpis: pd.DataFrame):
    ws = wb.create_sheet("Monthly Trends")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    ws["A1"].value = "Monthly Trends"
    ws["A1"].font = Font(bold=True, color=C_WHITE, size=14)
    ws["A1"].fill = _header_fill(C_DARK_BLUE)
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 28

    display = monthly_kpis[["order_month", "po_count", "total_spend", "total_delay_cost", "otd_rate"]].copy()
    display["order_month"] = display["order_month"].astype(str)
    _write_df_to_sheet(ws, display, start_row=3, header_color=C_MID_BLUE)

    data_end = 3 + len(display)

    # Combo chart: bar (spend) + line (OTD rate)
    bar = BarChart()
    bar.type = "col"
    bar.title = "Monthly Spend & On-Time Delivery Rate"
    bar.style = 10
    bar.width = 26
    bar.height = 14
    bar.y_axis.title = "Spend (USD)"
    bar.y_axis.axId = 100
    bar.x_axis.title = "Month"

    spend_col = display.columns.get_loc("total_spend") + 1
    otd_col   = display.columns.get_loc("otd_rate") + 1
    month_col = 1

    spend_ref = Reference(ws, min_col=spend_col, min_row=3, max_row=data_end)
    bar.add_data(spend_ref, titles_from_data=True)

    line = LineChart()
    line.title = "OTD Rate"
    line.style = 10
    line.y_axis.axId = 200
    line.y_axis.title = "OTD Rate"
    line.y_axis.crosses = "max"

    otd_ref = Reference(ws, min_col=otd_col, min_row=3, max_row=data_end)
    line.add_data(otd_ref, titles_from_data=True)

    bar += line
    cats = Reference(ws, min_col=month_col, min_row=4, max_row=data_end)
    bar.set_categories(cats)
    ws.add_chart(bar, "A" + str(data_end + 3))

    # Line chart: delay cost trend
    delay_line = LineChart()
    delay_line.title = "Monthly Delay Cost Trend"
    delay_line.style = 10
    delay_line.width = 22
    delay_line.height = 10
    delay_line.y_axis.title = "Delay Cost (USD)"

    delay_col = display.columns.get_loc("total_delay_cost") + 1
    delay_ref = Reference(ws, min_col=delay_col, min_row=3, max_row=data_end)
    delay_line.add_data(delay_ref, titles_from_data=True)
    delay_line.set_categories(cats)
    ws.add_chart(delay_line, "L" + str(data_end + 3))

    _auto_col_width(ws)
    ws.freeze_panes = "A4"


# ── Sheet 4: At-Risk Open Orders ──────────────────────────────────────────────

def _sheet_at_risk(wb: Workbook, at_risk: pd.DataFrame):
    ws = wb.create_sheet("At-Risk Open Orders")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:J1")
    ws["A1"].value = "At-Risk Open Orders — Requires Buyer Action"
    ws["A1"].font = Font(bold=True, color=C_WHITE, size=14)
    ws["A1"].fill = _header_fill(C_DARK_RED)
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 28

    if at_risk.empty:
        ws["A3"].value = "No at-risk orders identified."
        return

    display_cols = [
        "po_id", "supplier_name", "tier", "order_date", "promised_delivery_date",
        "days_since_order", "days_past_promised", "risk_level",
        "total_po_value", "risk_reason",
    ]
    # Only include columns that exist
    display_cols = [c for c in display_cols if c in at_risk.columns]
    display = at_risk[display_cols].copy()

    for date_col in ("order_date", "promised_delivery_date"):
        if date_col in display.columns:
            display[date_col] = display[date_col].dt.strftime("%Y-%m-%d")

    _write_df_to_sheet(ws, display, start_row=3, header_color=C_DARK_RED)

    # Colour each row by risk_level
    risk_col_idx = display_cols.index("risk_level") + 1
    for row_idx in range(4, 4 + len(display)):
        risk_cell = ws.cell(row=row_idx, column=risk_col_idx)
        level = risk_cell.value
        if level in RISK_FILLS:
            risk_cell.fill = RISK_FILLS[level]
            risk_cell.font = RISK_FONTS[level]
            risk_cell.alignment = _center()

    _auto_col_width(ws)
    ws.freeze_panes = "A4"


# ── Sheet 5: Raw Data ─────────────────────────────────────────────────────────

def _sheet_raw_data(wb: Workbook, orders: pd.DataFrame):
    ws = wb.create_sheet("Raw Data")

    raw = orders.copy()
    for col in ("order_date", "promised_delivery_date", "actual_delivery_date"):
        if col in raw.columns:
            raw[col] = raw[col].dt.strftime("%Y-%m-%d").where(raw[col].notna(), other=None)
    if "order_month" in raw.columns:
        raw["order_month"] = raw["order_month"].astype(str)

    _write_df_to_sheet(ws, raw, start_row=1, header_color=C_DARK_BLUE)
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    _auto_col_width(ws)


# ── Public entry point ────────────────────────────────────────────────────────

def generate_report(
    pipeline_results: dict,
    at_risk: pd.DataFrame,
    timestamp: str,
) -> str:
    """
    Build the full Excel workbook and write it to data/output/.
    Returns the file path.
    """
    os.makedirs(config.OUTPUT_DIR, exist_ok=True)

    wb = Workbook()
    # Remove default empty sheet
    del wb[wb.sheetnames[0]]

    _sheet_executive_summary(wb, pipeline_results, at_risk)
    _sheet_supplier_performance(wb, pipeline_results["supplier_kpis"])
    _sheet_monthly_trends(wb, pipeline_results["monthly_kpis"])
    _sheet_at_risk(wb, at_risk)
    _sheet_raw_data(wb, pipeline_results["orders"])

    path = os.path.join(config.OUTPUT_DIR, f"PO_Analytics_Report_{timestamp}.xlsx")
    wb.save(path)
    return path
